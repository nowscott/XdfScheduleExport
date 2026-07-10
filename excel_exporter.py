"""将课表导出为明细表和每日固定五时段的月视图。"""
import calendar
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import config


def _auto_fit_columns(ws, min_width: int = 10, max_width: int = 40):
    """自动调整列宽"""
    for col_idx in range(1, ws.max_column + 1):
        max_len = min_width
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                # 中文字符算 2 个宽度
                val = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, length)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, max_width)


# ---- 通用字段映射（自动识别常见字段名） ----

FIELD_NAME_MAP = {
    # 课程名
    "courseName": "课程名称", "coursename": "课程名称", "name": "课程名称",
    "className": "课程名称", "classname": "课程名称",
    "course_name": "课程名称", "lessonName": "课程名称",
    # 日期
    "date": "日期", "courseDate": "日期", "scheduleDate": "日期",
    "startDate": "日期", "day": "日期",
    # 时间
    "startTime": "开始时间", "endTime": "结束时间",
    "time": "时间", "courseTime": "时间",
    "beginTime": "开始时间", "finishTime": "结束时间",
    "startTimeStr": "开始时间", "endTimeStr": "结束时间",
    "lessonStartTime": "开始时间", "lessonEndTime": "结束时间",
    # 教室
    "room": "教室", "classRoom": "教室", "classroom": "教室",
    "roomName": "教室", "address": "地点",
    # 老师
    "teacher": "老师", "teacherName": "老师",
    "instructor": "老师", "lecturer": "老师",
    # 学生
    "student": "学生", "studentName": "学生",
    "lessonName": "学生",  # API 中 lessonName 实际是学生名
    "studentList": "学生名单",
    # 状态
    "status": "状态", "courseStatus": "状态",
    "lessonStatus": "状态",
    # 校区
    "campus": "校区", "schoolName": "校区",
    "school": "校区", "branch": "校区",
    # 备注
    "remark": "备注", "note": "备注", "memo": "备注",
    "description": "备注", "desc": "备注",
    # 年级/科目
    "grade": "年级", "subject": "科目",
    # 新 API 特有字段
    "lessonTypeDesc": "课次类型", "lessonType": "课次类型码",
    "classCode": "班级代码", "lessonId": "课次ID",
    "stuCount": "学生数", "schoolId": "校区ID",
    "resourceCount": "资源数", "videoType": "视频类型",
    "feedbackAllFinished": "反馈完成", "bindStatus": "绑定状态",
    "teachingChannelCodeList": "教学渠道",
    "_date": "日期",
}


def _map_field_name(raw: str) -> str:
    """将原始字段名映射为中文表头"""
    return FIELD_NAME_MAP.get(raw, raw)


# ---- 列排序优先级 ----
COLUMN_PRIORITY = [
    "日期", "开始时间", "结束时间", "学生", "课程名称",
    "老师", "教室", "校区", "课次类型", "状态", "备注",
]


def _sort_columns(headers: list[str]) -> list[str]:
    """按优先级排序列"""
    priority_map = {h: i for i, h in enumerate(COLUMN_PRIORITY)}
    # 未在优先级列表中的放最后
    return sorted(headers, key=lambda h: priority_map.get(h, 999))


# ---- 样式定义 ----

HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_FONT = Font(name="Microsoft YaHei", size=10)
CELL_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)

# 交替行颜色
EVEN_ROW_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

MONTH_TITLE_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
MONTH_WEEKDAY_FILL = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
MONTH_DATE_FILL = PatternFill(start_color="D9E7F5", end_color="D9E7F5", fill_type="solid")
MONTH_ACTIVE_DATE_FILL = PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid")
MONTH_COURSE_FILL = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
MONTH_WEEKEND_FILL = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")
MONTH_OUTSIDE_FILL = PatternFill(start_color="F5F6F8", end_color="F5F6F8", fill_type="solid")

EMPTY_SLOT_LABELS = (
    "08:00–10:00  无课",
    "10:00/10:20 时段  无课",
    "13:40–15:40  无课",
    "16:00–18:00  无课",
    "18:00/18:30 时段  无课",
)


def export_to_excel(schedules: list[dict],
                    start_date: str = "",
                    end_date: str = "",
                    output_path: str = None) -> str:
    """
    将课表数据导出为 Excel 文件

    Args:
        schedules: 课表数据列表
        start_date: 开始日期（用于文件名）
        end_date: 结束日期（用于文件名）
        output_path: 输出路径，不传则自动生成

    Returns:
        输出文件路径
    """
    if not schedules:
        raise ValueError("没有数据可导出")

    schedules = sorted(
        schedules,
        key=lambda item: (item.get("_date", ""), item.get("lessonStartTime", "")),
    )

    # ---- 解析字段 ----
    all_keys = list(schedules[0].keys())
    headers = [_map_field_name(k) for k in all_keys]
    # 创建原始字段名到中文名的映射
    raw_to_cn = dict(zip(all_keys, headers))

    # 去重但保持顺序
    seen = set()
    unique_headers = []
    for h in headers:
        if h not in seen:
            seen.add(h)
            unique_headers.append(h)

    # 按优先级排序
    sorted_headers = _sort_columns(unique_headers)

    # 反向映射：中文名 → 原始字段名
    cn_to_raw = {v: k for k, v in raw_to_cn.items()}

    # ---- 创建工作簿 ----
    wb = Workbook()
    ws = wb.active
    ws.title = "课表"

    # ---- 写表头 ----
    for col_idx, header in enumerate(sorted_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # ---- 写数据 ----
    for row_idx, item in enumerate(schedules, 2):
        for col_idx, header in enumerate(sorted_headers, 1):
            raw_key = cn_to_raw.get(header)
            value = item.get(raw_key, "") if raw_key else ""

            # 格式化特殊值
            if isinstance(value, dict):
                value = value.get("name", str(value))
            if isinstance(value, list):
                # studentList: [{studentName: ...}, ...] → 提取名字
                if value and isinstance(value[0], dict) and "studentName" in value[0]:
                    value = ", ".join(s.get("studentName", "") for s in value)
                else:
                    value = ", ".join(str(v) for v in value)
            # 格式化 lessonStatus 数字为可读状态
            if header == "状态" and isinstance(value, (int, str)):
                STATUS_MAP = {0: "未开始", 1: "进行中", 2: "已结束", "0": "未开始", "1": "进行中", "2": "已结束"}
                value = STATUS_MAP.get(value, value)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

            # 交替行背景色
            if row_idx % 2 == 0:
                cell.fill = EVEN_ROW_FILL

    # ---- 冻结首行 ----
    ws.freeze_panes = "A2"

    # ---- 自动筛选 ----
    ws.auto_filter.ref = f"A1:{get_column_letter(len(sorted_headers))}1"

    # ---- 自动列宽 ----
    _auto_fit_columns(ws)

    # ---- 生成月视图 ----
    _add_month_views(wb, schedules)

    # ---- 生成输出路径 ----
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if start_date and end_date:
            filename = f"课表_{start_date}_{end_date}_{timestamp}.xlsx"
        else:
            filename = f"课表_{timestamp}.xlsx"
        output_path = str(config.EXPORT_DIR / filename)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    # ---- 保存 ----
    wb.save(output_path)
    print(f"\n📁 已导出到: {output_path}")
    return output_path


def _get_student_name(lesson: dict) -> str:
    """从接口字段中提取学生姓名。"""
    if lesson.get("lessonName"):
        return str(lesson["lessonName"])
    students = lesson.get("studentList") or []
    if students and isinstance(students[0], dict):
        return "、".join(str(item.get("studentName", "")) for item in students if item.get("studentName"))
    return ""


def _time_part(value: object) -> str:
    text = str(value or "")
    match = re.search(r"(\d{2}:\d{2})", text)
    return match.group(1) if match else ""


def _short_room(value: object) -> str:
    room = str(value or "")
    match = re.search(r"个性化([^（(]+)", room)
    return match.group(1).strip() if match else room.replace("万博敏捷广场", "")


def _slot_index(start_time: str) -> int:
    hour, minute = (int(part) for part in start_time.split(":"))
    total = hour * 60 + minute
    if total < 9 * 60 + 30:
        return 0
    if total < 12 * 60 + 30:
        return 1
    if total < 15 * 60:
        return 2
    if total < 17 * 60 + 30:
        return 3
    return 4


def _add_month_views(wb: Workbook, schedules: list[dict]) -> None:
    grouped: dict[tuple[int, int], list[dict]] = defaultdict(list)
    for lesson in schedules:
        date_text = str(lesson.get("_date", ""))[:10]
        try:
            lesson_date = datetime.strptime(date_text, "%Y-%m-%d").date()
        except ValueError:
            continue
        grouped[(lesson_date.year, lesson_date.month)].append(lesson)

    for (year, month), month_lessons in sorted(grouped.items()):
        _add_month_sheet(wb, year, month, month_lessons)


def _add_month_sheet(wb: Workbook, year: int, month: int, lessons: list[dict]) -> None:
    title = f"{year}年{month}月月视图"
    ws = wb.create_sheet(title=title[:31])
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    by_date: dict[str, list[dict]] = defaultdict(list)
    for lesson in lessons:
        by_date[str(lesson.get("_date", ""))[:10]].append(lesson)
    for date_lessons in by_date.values():
        date_lessons.sort(key=lambda item: item.get("lessonStartTime", ""))

    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"{year} 年 {month} 月课程月视图"
    title_cell.fill = MONTH_TITLE_FILL
    title_cell.font = Font(name="Microsoft YaHei", size=20, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:G2")
    summary_cell = ws["A2"]
    summary_cell.value = (
        f"共 {len(lessons)} 节课 · {len(by_date)} 个有课日期 · "
        "每个日期固定 5 个时段，一节课占一行"
    )
    summary_cell.fill = PatternFill(start_color="D9E7F5", end_color="D9E7F5", fill_type="solid")
    summary_cell.font = Font(name="Microsoft YaHei", size=10, color="365F91")
    summary_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    weekdays = ("星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日")
    for column, label in enumerate(weekdays, 1):
        cell = ws.cell(3, column, label)
        cell.fill = MONTH_WEEKDAY_FILL
        cell.font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[3].height = 28

    month_calendar = calendar.Calendar(firstweekday=calendar.MONDAY).monthdayscalendar(year, month)
    for week_index, week in enumerate(month_calendar):
        date_row = 4 + week_index * 6
        ws.row_dimensions[date_row].height = 22

        for weekday_index, day in enumerate(week):
            column = weekday_index + 1
            date_cell = ws.cell(date_row, column)
            date_cell.border = THIN_BORDER
            date_cell.alignment = Alignment(horizontal="left", vertical="center")
            date_cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="365F91")

            if day == 0:
                date_cell.fill = MONTH_OUTSIDE_FILL
                for slot in range(5):
                    empty_cell = ws.cell(date_row + slot + 1, column)
                    empty_cell.fill = MONTH_OUTSIDE_FILL
                    empty_cell.border = THIN_BORDER
                continue

            date_key = f"{year:04d}-{month:02d}-{day:02d}"
            day_lessons = by_date.get(date_key, [])
            date_cell.value = f"{day} 日" + (f" · {len(day_lessons)} 节" if day_lessons else "")
            date_cell.fill = MONTH_ACTIVE_DATE_FILL if day_lessons else (
                MONTH_DATE_FILL if weekday_index < 5 else PatternFill(
                    start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"
                )
            )

            slots: list[dict | None] = [None] * 5
            for lesson in day_lessons:
                start = _time_part(lesson.get("lessonStartTime"))
                if not start:
                    continue
                index = _slot_index(start)
                if slots[index] is not None:
                    raise ValueError(f"{date_key} 的第 {index + 1} 时段存在多节课程，无法放入固定五行月视图")
                slots[index] = lesson

            for slot, lesson in enumerate(slots):
                row = date_row + slot + 1
                cell = ws.cell(row, column)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="left", vertical="center")
                ws.row_dimensions[row].height = 23

                if lesson is None:
                    cell.value = EMPTY_SLOT_LABELS[slot]
                    cell.fill = MONTH_WEEKEND_FILL if weekday_index >= 5 else PatternFill(
                        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
                    )
                    cell.font = Font(name="Microsoft YaHei", size=8, color="7F8C9A")
                    continue

                start = _time_part(lesson.get("lessonStartTime"))
                end = _time_part(lesson.get("lessonEndTime"))
                student = _get_student_name(lesson)
                room = _short_room(lesson.get("roomName"))
                cell.value = f"{start}–{end}  {student} · {room}"
                cell.fill = MONTH_COURSE_FILL
                cell.font = Font(name="Microsoft YaHei", size=8, bold=True, color="375623")

    for column in range(1, 8):
        ws.column_dimensions[get_column_letter(column)].width = 27

    last_row = 3 + len(month_calendar) * 6
    ws.print_area = f"A1:G{last_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
