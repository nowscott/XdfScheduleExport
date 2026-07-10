import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from excel_exporter import export_to_excel


class ExportTest(unittest.TestCase):
    def test_exports_detail_and_fixed_five_slot_month_view(self):
        schedules = [
            {
                "_date": "2026-07-13",
                "lessonStartTime": "2026-07-13 10:20:00",
                "lessonEndTime": "2026-07-13 12:20:00",
                "lessonName": "王同学",
                "teacherName": "牛老师",
                "roomName": "万博敏捷广场个性化V229（G）",
                "lessonStatus": 0,
            },
            {
                "_date": "2026-07-13",
                "lessonStartTime": "2026-07-13 16:00:00",
                "lessonEndTime": "2026-07-13 18:00:00",
                "lessonName": "李同学",
                "teacherName": "牛老师",
                "roomName": "万博敏捷广场个性化V230（G）",
                "lessonStatus": 0,
            },
        ]

        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "schedule.xlsx"
            export_to_excel(schedules, "2026-07-13", "2026-07-31", str(output))

            workbook = load_workbook(output)
            self.assertEqual(workbook.sheetnames, ["课表", "2026年7月月视图"])

            month = workbook["2026年7月月视图"]
            self.assertEqual(month["A16"].value, "13 日 · 2 节")
            self.assertIn("08:00–10:00", month["A17"].value)
            self.assertIn("无课", month["A17"].value)
            self.assertIn("10:20–12:20", month["A18"].value)
            self.assertIn("王同学", month["A18"].value)
            self.assertIn("16:00–18:00", month["A20"].value)
            self.assertIn("李同学", month["A20"].value)
            self.assertIn("无课", month["A21"].value)


if __name__ == "__main__":
    unittest.main()

